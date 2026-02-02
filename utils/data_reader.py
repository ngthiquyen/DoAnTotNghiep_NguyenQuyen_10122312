import os
import pandas as pd
import json
import numpy as np
from openpyxl import load_workbook


def load_test_data(file_path, sheet_name=None):
    """Đọc test data từ Excel / CSV / JSON"""
    ext = os.path.splitext(file_path)[-1].lower()

    if ext in [".xlsx", ".xls"]:
        return read_excel(file_path, sheet_name)
    elif ext == ".csv":
        return read_csv(file_path)
    elif ext == ".json":
        return read_json(file_path)
    else:
        raise ValueError(f"Định dạng file không được hỗ trợ: {ext}")


def read_csv(file_path):
    df = pd.read_csv(file_path, dtype=str)
    df = df.replace({np.nan: ""})
    return df.to_dict(orient="records")


def read_excel(file_path, sheet_name=0):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[sheet_name] if isinstance(sheet_name, int) else sheet_name]

    data = list(ws.values)
    headers = [str(h).strip() for h in data[0]]

    records = []
    for row in data[1:]:
        record = {}
        for i, cell in enumerate(row):
            record[headers[i]] = "" if cell is None else str(cell).strip()
        records.append(record)

    return records


def read_json(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)
